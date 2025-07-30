import os
import sys
import json
from docx import Document
from openpyxl import load_workbook
from colorama import Fore, Style, init

# Įtraukiam core modulį iš submodule
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'MetaCore_FIRMWARE')))
from core.index_manager import regenerate_firmware_index

# 🎨 Spalvų inicializavimas
init(autoreset=True)

# === Konfigūracijos ===
FIRMWARE_DIR = "MetaCore_FIRMWARE/core"
INDEX_PATH = "MetaCore_FIRMWARE/config/firmware_index.json"

# === DOCX META HEADERS ===
def read_docx_meta(path):
    try:
        doc = Document(path)
        text = "\n".join(p.text.strip() for p in doc.paragraphs if p.text.strip())
        if "#VTXT_META_HEADER_START" in text:
            lines = text.split("#VTXT_META_HEADER_START")[1].splitlines()
            meta = {}
            for line in lines:
                if "]: " in line:
                    try:
                        key, value = line.strip().strip("[]").split("]: ", 1)
                        meta[key.strip()] = value.strip().strip('"“”')
                    except ValueError:
                        continue
            return meta
        return {"info": text[:300]}
    except Exception as e:
        return {"error": f"❌ Error reading DOCX: {e}"}

# === XLSX APŽVALGA ===
def read_xlsx_preview(path):
    try:
        wb = load_workbook(filename=path, data_only=True)
        preview = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            preview.append(f"🔹 Sheet: {sheet}")
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                preview.append("   " + str(row))
        return "\n".join(preview)
    except Exception as e:
        return f"❌ Error reading XLSX: {e}"

# === Failų sąrašas ===
def list_firmwares():
    return [
        f for f in os.listdir(FIRMWARE_DIR)
        if f.startswith("firmware_") and f.endswith((".docx", ".xlsx"))
    ]

# === Įkelti JSON indeksą ===
def load_index():
    try:
        with open(INDEX_PATH, "r") as f:
            return json.load(f)
    except Exception as e:
        print(Fore.RED + f"⚠️ Unable to load index: {e}")
        return []

# === Rodyti vieną failą ===
def display_firmware(file):
    path = os.path.join(FIRMWARE_DIR, file)
    print(f"\n{Fore.CYAN}📦 {file}")
    print(f"{Style.DIM}{'-'*60}")
    if file.endswith(".docx"):
        meta = read_docx_meta(path)
        for k, v in meta.items():
            print(f"{Fore.YELLOW}{k}: {Fore.WHITE}{v}")
    elif file.endswith(".xlsx"):
        print(read_xlsx_preview(path))

# === Pagrindinė programa ===
def main():
    print(Fore.GREEN + "🌐 Scanning FIRMWARE Modules...\n")
    fw_list = list_firmwares()
    if not fw_list:
        print(Fore.RED + "❌ No firmware files found.")
        return

    index_data = load_index()
    registered = {
        mod["file"]: mod
        for group in index_data
        for mod in group.get("modules", [])
    }

    for fw in fw_list:
        display_firmware(fw)
        fw_key = fw.replace(".docx", "").replace(".xlsx", "")
        if fw_key in registered:
            print(Fore.GREEN + "✔ Indexed in firmware_index.json")
        else:
            print(Fore.RED + "✖ Not found in firmware_index.json")

# === Argumentai ===
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--refresh", action="store_true", help="Regenerate firmware index before listing")
    args = parser.parse_args()

    if args.refresh:
        print(Fore.MAGENTA + "🔄 Regenerating firmware index...")
        regenerate_firmware_index()

    main()
